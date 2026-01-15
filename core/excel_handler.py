# -*- coding: utf-8 -*-
"""
Excel Handler Module - Xử lý đọc/ghi file Excel
"""

import os
import logging
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config.settings import (
    EXCEL_TEMPLATE_COLUMNS, DATA_DIR,
    DEFAULT_TYPE, DEFAULT_ASPECT_RATIO, DEFAULT_DURATION,
    DEFAULT_RESOLUTION, DEFAULT_VARIATIONS
)

logger = logging.getLogger(__name__)


@dataclass
class TaskRow:
    """Đại diện cho một hàng task trong Excel"""
    row_number: int
    prompt: str
    image_path: str = ""  # Đường dẫn ảnh để upload
    type: str = DEFAULT_TYPE
    aspect_ratio: str = DEFAULT_ASPECT_RATIO
    duration: str = DEFAULT_DURATION
    resolution: str = DEFAULT_RESOLUTION
    variations: int = DEFAULT_VARIATIONS
    output_path: str = ""
    status: str = ""
    result: str = ""
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "row_number": self.row_number,
            "prompt": self.prompt,
            "image_path": self.image_path,
            "type": self.type,
            "aspect_ratio": self.aspect_ratio,
            "duration": self.duration,
            "resolution": self.resolution,
            "variations": self.variations,
            "output_path": self.output_path,
            "status": self.status,
            "result": self.result
        }


class ExcelHandler:
    """Xử lý đọc/ghi file Excel"""
    
    def __init__(self, filepath: str = None):
        """
        Khởi tạo handler
        
        Args:
            filepath: Đường dẫn file Excel
        """
        self.filepath = filepath or os.path.join(DATA_DIR, "sora.xlsx")
        self.workbook: Optional[Workbook] = None
        self.sheet = None
    
    def create_template(self, filepath: str = None) -> str:
        """
        Tạo file template Excel
        
        Args:
            filepath: Đường dẫn lưu file
            
        Returns:
            Đường dẫn file đã tạo
        """
        filepath = filepath or self.filepath
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        
        # Style cho header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Ghi header
        for col, header in enumerate(EXCEL_TEMPLATE_COLUMNS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Thiết lập độ rộng cột
        column_widths = {
            'A': 50,  # Prompt
            'B': 40,  # ImagePath
            'C': 10,  # Type
            'D': 12,  # AspectRatio
            'E': 10,  # Duration
            'F': 12,  # Resolution
            'G': 12,  # Variations
            'H': 40,  # OutputPath
            'I': 15,  # Status
            'J': 30,  # Result
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Thêm dòng mẫu
        sample_data = [
            "A beautiful sunset over the ocean with golden light",
            "",  # ImagePath (để trống hoặc điền đường dẫn ảnh)
            "video",
            "3:2",
            "10s",
            "720p",
            1,
            "",
            "Pending",
            ""
        ]
        
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border
        
        # Lưu file
        os.makedirs(os.path.dirname(filepath) if os.path.dirname(filepath) else ".", exist_ok=True)
        wb.save(filepath)
        logger.info(f"Đã tạo template: {filepath}")
        
        return filepath
    
    def load(self, filepath: str = None) -> bool:
        """
        Load file Excel
        
        Args:
            filepath: Đường dẫn file
            
        Returns:
            True nếu load thành công
        """
        filepath = filepath or self.filepath
        
        if not os.path.exists(filepath):
            logger.warning(f"File không tồn tại: {filepath}")
            return False
        
        try:
            self.workbook = load_workbook(filepath)
            self.sheet = self.workbook.active
            self.filepath = filepath
            logger.info(f"Đã load file: {filepath}")
            return True
        except Exception as e:
            logger.error(f"Lỗi load file Excel: {e}")
            return False
    
    def get_tasks(self, include_completed: bool = False) -> List[TaskRow]:
        """
        Lấy danh sách tasks từ Excel
        
        Args:
            include_completed: Có bao gồm task đã hoàn thành không
            
        Returns:
            Danh sách TaskRow
        """
        if not self.sheet:
            logger.error("Chưa load file Excel")
            return []
        
        tasks = []
        
        for row_idx in range(2, self.sheet.max_row + 1):
            prompt = self.sheet.cell(row=row_idx, column=1).value
            
            if not prompt or str(prompt).strip() == "":
                continue
            
            status = str(self.sheet.cell(row=row_idx, column=9).value or "").strip()
            
            # Bỏ qua task đã hoàn thành nếu không cần
            if not include_completed and status.lower() in ["completed", "done", "success", "hoàn thành"]:
                continue
            
            task = TaskRow(
                row_number=row_idx,
                prompt=str(prompt).strip(),
                image_path=str(self.sheet.cell(row=row_idx, column=2).value or "").strip(),
                type=str(self.sheet.cell(row=row_idx, column=3).value or DEFAULT_TYPE).strip().lower(),
                aspect_ratio=str(self.sheet.cell(row=row_idx, column=4).value or DEFAULT_ASPECT_RATIO).strip(),
                duration=str(self.sheet.cell(row=row_idx, column=5).value or DEFAULT_DURATION).strip(),
                resolution=str(self.sheet.cell(row=row_idx, column=6).value or DEFAULT_RESOLUTION).strip(),
                variations=int(self.sheet.cell(row=row_idx, column=7).value or DEFAULT_VARIATIONS),
                output_path=str(self.sheet.cell(row=row_idx, column=8).value or "").strip(),
                status=status,
                result=str(self.sheet.cell(row=row_idx, column=10).value or "").strip()
            )
            
            tasks.append(task)
        
        logger.info(f"Tìm thấy {len(tasks)} task(s)")
        return tasks
    
    def update_status(self, row_number: int, status: str, result: str = None, save: bool = True):
        """
        Cập nhật trạng thái task
        
        Args:
            row_number: Số dòng trong Excel
            status: Trạng thái mới
            result: Kết quả (nếu có)
            save: Có lưu file không
        """
        if not self.sheet:
            logger.error("Chưa load file Excel")
            return
        
        self.sheet.cell(row=row_number, column=9, value=status)
        
        if result is not None:
            self.sheet.cell(row=row_number, column=10, value=result)
        
        if save:
            self.save()
        
        logger.info(f"Đã cập nhật dòng {row_number}: {status}")
    
    def update_output_path(self, row_number: int, output_path: str, save: bool = True):
        """Cập nhật đường dẫn output"""
        if not self.sheet:
            return
        
        self.sheet.cell(row=row_number, column=8, value=output_path)
        
        if save:
            self.save()
    
    def save(self, filepath: str = None):
        """Lưu file Excel"""
        if not self.workbook:
            return
        
        filepath = filepath or self.filepath
        
        try:
            self.workbook.save(filepath)
            logger.info(f"Đã lưu file: {filepath}")
        except Exception as e:
            logger.error(f"Lỗi lưu file: {e}")
    
    def close(self):
        """Đóng workbook"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
            self.sheet = None
